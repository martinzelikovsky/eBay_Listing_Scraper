from selenium import webdriver
import requests
import openpyxl
import time
import os
import fnmatch

class eBay_Scraper:
       '''
       This web scraper is for getting your listing images from your listings. Creates a top level folder with children
       folders containing listing images per listing and a copy of the listing CSV file with an added column linking
       each listing its image directory. Each listing link creates a folder with the listing images.

       Uses Chrome Driver for Version 88 of Google Chrome. Link to download:
       https://chromedriver.storage.googleapis.com/index.html?path=88.0.4324.96/
       '''

       def __init__(self, folder, result_name, webdriver_path):
              self.cwd = os.getcwd()
              self.input_folder = folder
              self.result_folder = os.path.join(os.getcwd(), '{}_Results'.format(result_name))
              self.sheet_names = fnmatch.filter(os.listdir(self.input_folder), '*.xlsx')

              self.chromedriver_link = 'https://chromedriver.storage.googleapis.com/index.html?path=88.0.4324.96/'
              self.driver = webdriver.Chrome(executable_path=webdriver_path)
              self.expand_image_id = 'icImg' # Button to expand image in listing
              self.image_id = 'viEnlargeImgLayer_img_ctr' # Image ID
              self.next_image_button_class = 'pntrArr.pntrArrNext.pntrArrImg.activeNext' # Button to select next image
              self.i = 0
              self.j = 0

       def get_worksheet(self, sheet_name):
              sheet_path = os.path.join(self.input_folder, sheet_name)
              self.wb = openpyxl.load_workbook(filename=sheet_path)
              self.ws = self.wb['Listings']
              self.ws_columns = list(self.ws.columns)
              self.IDs = []
              self.titles = []
              for row in range(4, len(self.ws_columns[0])):
                     self.IDs.append(self.ws_columns[1][row].value)
                     self.titles.append(self.ws_columns[3][row].value.replace(' ', '_').replace('/', '_').replace('\\', '-').replace('"', '').replace("'",''))

       def get_result_directories(self):
              # Main result folder
              if not os.path.exists(self.result_folder):
                     os.mkdir(self.result_folder)
              # Listing Worksheet folder
              sheet_folder = self.sheet_names[self.i].strip('.xlsx')
              self.sheet_folder_path = os.path.join(self.result_folder, sheet_folder)
              if not os.path.exists(self.sheet_folder_path):
                     os.mkdir(self.sheet_folder_path)
              # Item images folder
              image_folder = self.titles[self.j]
              self.image_folder_path = os.path.join(self.sheet_folder_path, image_folder)
              if not os.path.exists(self.image_folder_path):
                     os.mkdir(self.image_folder_path)

       def get_images(self):
              if self.j == 0:
                     self.folder_col = len(self.ws_columns) + 1
                     self.ws.cell(4, self.folder_col).value = 'Image Folder Path'

              listing_url = 'https://www.ebay.com/itm/{}'.format(self.IDs[self.j])
              self.ws.cell(self.j + 5, self.folder_col).value = '=HYPERLINK("{}")'.format(self.image_folder_path)
              self.driver.get(listing_url)
              try:
                     self.driver.find_element_by_id(self.expand_image_id).click() # Clicks to expand image
              except:
                     return
              image_num = 1
              while True:
                     time.sleep(0.5)
                     image_link = self.driver.find_element_by_id(self.image_id).get_attribute('src') # Finds image source link
                     image_name = os.path.join(self.image_folder_path, 'Image_{}'.format(image_num) + '.jpg')
                     with open(image_name, 'wb') as pic: # Downloads image
                            im = requests.get(image_link)
                            pic.write(im.content)
                     image_num += 1
                     try:
                            self.driver.find_element_by_class_name(self.next_image_button_class).click() # Clicks to go to next image
                     except:
                            print('Finshed listing "{}"'.format(self.titles[self.j])) # Listing complete
                            break

       def run(self):
              for i, name in enumerate(self.sheet_names):
                     self.i = i
                     self.get_worksheet(name)

                     for j, title in enumerate(self.titles):
                            self.j = j
                            self.get_result_directories()
                            self.get_images()

                     workbook_path = os.path.join(self.sheet_folder_path, 'Listings.xlsx')
                     self.wb.save(workbook_path)
              self.driver.close()


if __name__ == '__main__':
       folder_path = input('Please enter full path of the parent folder containing all your eBay CSV Listing Exports: ')
       result_name = input('Please enter desired name of result folder for this listing scraping session (folder will be'
                           ' created in the same directory as this file exists in): ')
       webdriver_path = input('Please enter full path of the Chrome Driver (typically: C:\\WebDriver\\bin\\chromedriver.exe)...')
       scrape = eBay_Scraper(folder=folder_path, result_name=result_name, webdriver_path=webdriver_path)
       scrape.run()




